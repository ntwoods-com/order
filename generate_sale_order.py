# ================================
# generate_sale_order.py (FINAL UPDATED)
# ================================
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from datetime import datetime
import os
import sqlite3
from db_utils import connect as db_connect, init_schema

def _get_db_file() -> str:
    return os.getenv("DATABASE_FILE", os.path.join(os.path.dirname(os.path.abspath(__file__)), "order_counter.db"))

# ======================================================
# GLOBAL STYLES / THEME (Revised to match the image + Previous Logic)
# ======================================================

# Borders
thin   = Side(border_style="thin",   color="000000")
thick  = Side(border_style="thick",  color="000000")
medium = Side(border_style="medium", color="000000")

border_all    = Border(top=thin, left=thin, right=thin, bottom=thin)
border_thick  = Border(top=thick, left=thick, right=thick, bottom=thick)
border_header = Border(top=medium, left=medium, right=medium, bottom=medium)

# Theme palette (hex) - Based on the Image (Screenshot)
CL_WHITE        = "FFFFFF"
CL_BLACK        = "000000"
CL_ACCENT_DARK  = "1B4D2A"  # Dark Green (Top Title/Order Info Header)
CL_ACCENT_MED   = "5E89A3"  # Medium Blue/Teal (Section Headers/Brand Total)
CL_ACCENT_LIGHT = "C9E1EF"  # Brand/Category Strips 
CL_GREEN_SUBT   = "6AA84F"  # Subtotal Green
CL_GOLD_TOTAL   = "FFC000"  # Grand Total Band

# Order Info Section Colors (Matching the image)
CL_INFO_HEADER  = "1B4D2A"  # Dark Green/Teal for ORDER INFORMATION title bar
CL_INFO_LABEL   = "D9D9D9"  # Light Gray for Label Column (A)
CL_INFO_VALUE   = "FFFFFF"  # White for Value Column (B-E merged)
CL_YELLOW_ID    = "FFF2CC"  # Light Yellow for ORDER ID

CL_TOP_DARK     = "1B4D2A"  # Topmost Dark Green Title
CL_FOOTER_GRAY  = "EDEDED"  # Footer

# Fills
FILL_WHITE       = PatternFill("solid", fgColor=CL_WHITE)
FILL_BLUE_HDR    = PatternFill("solid", fgColor=CL_ACCENT_MED)    # Section Header
FILL_GREEN_DARK  = PatternFill("solid", fgColor=CL_ACCENT_DARK)   # Table Header / Brand Total
FILL_ACCENT_LIGHT     = PatternFill("solid", fgColor=CL_ACCENT_LIGHT)  # Brand/Category Strip
FILL_GOLD        = PatternFill("solid", fgColor=CL_GOLD_TOTAL)
FILL_GRAY_MED    = PatternFill("solid", fgColor=CL_INFO_LABEL)    # Info Label Column
FILL_GRAY_LIGHT  = PatternFill("solid", fgColor="E6E7EB")         # Zebra rows (subtler light gray)
FILL_YELLOW_ID   = PatternFill("solid", fgColor=CL_YELLOW_ID)
FILL_TOP_DARK    = PatternFill("solid", fgColor=CL_TOP_DARK)      # Top Title Bar
FILL_FOOTER      = PatternFill("solid", fgColor=CL_FOOTER_GRAY)
FILL_INFO_HEADER = PatternFill("solid", fgColor=CL_INFO_HEADER)   # ORDER INFORMATION Header

# --------------- Generic row styling helper ---------------
def style_row(ws, row_num, *, start_col=1, end_col=7, bold=False, size=10,
              text_color=CL_BLACK, fill=None, border=border_all,
              halign="center", valign="center", wrap=False):
    """Apply uniform font/fill/border/alignment to a row segment."""
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font = Font(bold=bold, size=size, name="Calibri", color=text_color)
        if fill:
            cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)

# =========================
# ORDER ID (SQLite counter)
# =========================
def generate_unique_order_id():
    """Unique ORDER ID: MM-YY-NNNNN (persists per month)."""
    db_file = _get_db_file()
    try:
        init_schema(default_sqlite_db_file=db_file)
        conn = db_connect(default_sqlite_db_file=db_file)

        mm_yy = datetime.now().strftime("%m-%y")
        row = conn.execute("SELECT counter FROM counters WHERE month_year=?", (mm_yy,)).fetchone()
        if row is None:
            new_counter = 1
            conn.execute("INSERT INTO counters (month_year, counter) VALUES (?, ?)", (mm_yy, new_counter))
        else:
            new_counter = int(row["counter"]) + 1
            conn.execute("UPDATE counters SET counter=? WHERE month_year=?", (new_counter, mm_yy))

        conn.commit()
        conn.close()
        return f"{mm_yy}-{str(new_counter).zfill(5)}"
    except Exception as e:
        print(f"[DB] Counter error: {e}")
        return "ERROR-ID"

def log_order_to_database(username, dealer_name, city, order_id, report_name):
    """Audit log for generated orders."""
    try:
        db_file = _get_db_file()
        init_schema(default_sqlite_db_file=db_file)
        conn = db_connect(default_sqlite_db_file=db_file)

        generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute(
            "INSERT INTO sale_orders (username, dealer_name, city, order_id, report_name, generated_at) VALUES (?, ?, ?, ?, ?, ?)",
            (username, dealer_name, city, order_id, report_name, generated_at),
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[DB] Log error: {e}")

# =========================
# DOMAIN HELPERS
# =========================
def extract_thickness_from_category(category):
    if category is None:
        return None
    m = re.match(r"(\d+\.?\d*)mm", str(category).strip().lower())
    return float(m.group(1)) if m else None

def get_sqft_formula(row, row_num):
    product = str(row["PRODUCT"]).lower()
    if product in ["laminate", "liner"]:
        return "0"
    # size in B, qty in E
    return (
        f'=LET(a,LEFT(B{row_num},1)*RIGHT(B{row_num},1)*E{row_num},'
        f'b,LEFT(B{row_num},2)*RIGHT(B{row_num},2)/144*E{row_num},'
        f'IF(LEN(B{row_num})<4,a,b))'
    )

def get_weight_formula(row, weight_map, hdmr_map, mdf_map, ply_map, pvc_map, wpc_map, row_num):
    product = str(row["PRODUCT"]).lower()
    brand   = str(row["BRAND"]).upper()
    category= str(row["CATEGORY"]).lower()

    hardcoded = {"door": 1.5, "board": 1}
    if product in hardcoded:
        return f"=F{row_num}*{hardcoded[product]}"

    thickness = extract_thickness_from_category(category)

    if product == "hdmr" and thickness in hdmr_map:
        return f"=E{row_num}*{hdmr_map[thickness]}"
    if product == "mdf" and thickness in mdf_map:
        return f"=E{row_num}*{mdf_map[thickness]}"
    if product == "ply" and thickness in ply_map:
        return f"=F{row_num}*{ply_map[thickness]}"           # ply uses SQFT
    if product == "pvc door" and thickness in pvc_map:
        return f"=F{row_num}*{pvc_map[thickness]}"           # pvc uses SQFT
    if product == "wpc board" and thickness in wpc_map:
        return f"=E{row_num}*{wpc_map[thickness]}"

    key = (product, brand)
    if product in ["laminate", "liner"] and key in weight_map:
        return f"=E{row_num}*{weight_map[key]}"

    return "0"

def normalize_category(raw_category, cat_map_df, product):
    """
    Normalizes category names based on a mapping table.
    """
    if pd.isna(raw_category):
        return None
    
    raw_cat = str(raw_category).upper()

    if "TEX" in raw_cat:
        return "TEX CATEGORY"

    product = str(product).lower()
    if product in ['laminate', 'liner']:
        for _, row in cat_map_df.iterrows():
            keyword = str(row['MATCH KEYWORD']).upper()
            target = row['NORMALIZED CATEGORY']
            if keyword == '*':
                continue
            elif '+' in keyword:
                if all(k.strip() in raw_cat for k in keyword.split('+')):
                    return target
            elif keyword in raw_cat:
                return target
        default = cat_map_df[cat_map_df['MATCH KEYWORD'] == '*']['NORMALIZED CATEGORY']
        return default.values[0] if not default.empty else raw_cat
    
    return raw_category

def calculate_sqft(size, qty):
    try:
        if not isinstance(size, str) or "X" not in size.upper():
            return 0
        l_str, b_str = re.split(r"[xX]", size)
        l, b = float(l_str), float(b_str)
        if l <= 15 and b <= 15:
            return l * b * qty
        return (l * b / 144) * qty
    except Exception:
        return 0

# =========================
# DATA PREPARATION
# =========================
def prepare_data(input_file):
    """Load excel, build maps, normalize categories, sort."""
    try:
        df = pd.read_excel(input_file, sheet_name="Master")
        cat_map = pd.read_excel(input_file, sheet_name="CategoryMap")
        weight_map_df = pd.read_excel(input_file, sheet_name="WeightMap")
        hdmr_map_df   = pd.read_excel(input_file, sheet_name="HDMRWeightMap")
        mdf_map_df    = pd.read_excel(input_file, sheet_name="MDFWeightMap")
        ply_map_df    = pd.read_excel(input_file, sheet_name="PlyWeightMap")
        pvc_map_df    = pd.read_excel(input_file, sheet_name="PVCWeightMap")
        wpc_map_df    = pd.read_excel(input_file, sheet_name="WPCBoardWeightMap")
    except Exception as e:
        print(f"[WARN] Excel read issue: {e} — using minimal fallback data.")
        df = pd.DataFrame({
            "PRODUCT": ["Door"],
            "SIZE": ["72x30"],
            "CATEGORY": ["-"],
            "BRAND": ["Default"],
            "QUANTITY": [3],
        })
        cat_map = pd.DataFrame({"MATCH KEYWORD": ["*"], "NORMALIZED CATEGORY": ["Default"]})
        weight_map_df = hdmr_map_df = mdf_map_df = ply_map_df = pvc_map_df = wpc_map_df = pd.DataFrame()

    # maps
    try:
        weight_map = {(str(r["PRODUCT"]).lower(), str(r["BRAND"]).upper()): r["WEIGHT_PER_PCS"]
                      for _, r in weight_map_df.iterrows()}
    except Exception:
        weight_map = {}

    def _tmap(df_map, col):
        try:
            return {float(r["THICKNESS"]): r[col] for _, r in df_map.iterrows()}
        except Exception:
            return {}

    hdmr_map = _tmap(hdmr_map_df, "WEIGHT_PER_PCS")
    mdf_map  = _tmap(mdf_map_df,  "WEIGHT_PER_PCS")
    ply_map  = _tmap(ply_map_df,  "WEIGHT_PER_SQFT")
    pvc_map  = _tmap(pvc_map_df,  "WEIGHT_PER_SQFT")
    wpc_map  = _tmap(wpc_map_df,  "WEIGHT_PER_PCS")

    # ************************************************
    #  UPDATED CATEGORY SORTING LOGIC 
    #  Order: SF -> HG -> CategoryMap Series -> Others
    # ************************************************
    try:
        mapped_categories = cat_map[cat_map["MATCH KEYWORD"]!="*"]["NORMALIZED CATEGORY"].drop_duplicates().tolist()
        df["CATEGORY_NORM"] = df.apply(lambda r: normalize_category(r["CATEGORY"], cat_map, r["PRODUCT"]), axis=1)
        df["CATEGORY_NORM"] = df["CATEGORY_NORM"].fillna("UNSPECIFIED")

        all_unique = df["CATEGORY_NORM"].unique().tolist()
        cat_order = []
        if "SF" in all_unique: cat_order.append("SF")
        if "HG" in all_unique: cat_order.append("HG")
        for c in mapped_categories:
            if c in all_unique and c not in ["SF","HG","TEX CATEGORY"]:
                cat_order.append(c)
        if "TEX CATEGORY" in all_unique: cat_order.append("TEX CATEGORY")
        if "UNSPECIFIED" in all_unique:  cat_order.append("UNSPECIFIED")
        for c in all_unique:
            if c not in cat_order: cat_order.append(c)
    except Exception:
        cat_order = []

    # cleaning
    for col in ["PRODUCT","SIZE","BRAND","CATEGORY"]:
        df[col] = df[col].astype(str).str.strip()
    df["QUANTITY"] = pd.to_numeric(df["QUANTITY"], errors="coerce").fillna(0)
    df["SQFT"] = df.apply(lambda r: calculate_sqft(r["SIZE"], r["QUANTITY"]), axis=1)

    # sorting
    try:
        df["SIZE_NUMERIC"] = pd.to_numeric(df["SIZE"], errors="coerce")
        df = df.sort_values(["BRAND","CATEGORY_NORM","PRODUCT","SIZE_NUMERIC"],
                            ascending=True, na_position="last").drop(columns=["SIZE_NUMERIC"])
    except Exception:
        df = df.sort_values(["BRAND","CATEGORY_NORM","PRODUCT","SIZE"], ascending=True)

    if cat_order:
        df["CATEGORY_NORM"] = pd.Categorical(df["CATEGORY_NORM"], categories=cat_order, ordered=True)
        df = df.sort_values(["BRAND","CATEGORY_NORM","PRODUCT","SIZE"], ascending=True)

    return df, cat_order, weight_map, hdmr_map, mdf_map, ply_map, pvc_map, wpc_map

# =========================
# REPORT GENERATION
# =========================
def write_report(
    df,
    output_file,
    weight_map, hdmr_map, mdf_map, ply_map, pvc_map, wpc_map,
    username, dealer_name, city, order_date, freight_condition
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SALE ORDER"

    # Column widths
    widths = {1: 22, 2: 14, 3: 16, 4: 18, 5: 10, 6: 12, 7: 12}
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---------- Top Bar Row 1: Dark Green title ----------
    for c in range(1, 8):
        ws.cell(row=1, column=c).fill = FILL_TOP_DARK
        ws.cell(row=1, column=c).border = border_all
    ws.merge_cells('A1:G1')
    t1 = ws['A1']; t1.value = "NT WOODS PVT.LTD"
    t1.font = Font(bold=True, size=14, color=CL_WHITE)
    t1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ---------- Row 2: Blue "PROVISIONAL SALE ORDER" ----------
    for c in range(1, 8):
        ws.cell(row=2, column=c).fill = FILL_BLUE_HDR
        ws.cell(row=2, column=c).border = border_all
    ws.merge_cells('A2:G2')
    t2 = ws['A2']; t2.value = "PROVISIONAL SALE ORDER"
    t2.font = Font(bold=True, size=16, color=CL_WHITE)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 26

    # ---------- Row 3: Subtitle left + Generated right ----------
    ws.merge_cells('A3:D3')
    ws['A3'].value = "N T WOOD PVT. LTD - Premium Wood Solutions"
    ws['A3'].font = Font(bold=True, size=11)
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('E3:G3')
    ws['E3'].value = f"Generated: {datetime.now().strftime('%d-%m-%Y-%H:%M')}"
    ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 8):
        ws.cell(row=3, column=c).border = border_all
        ws.cell(row=3, column=c).fill = FILL_WHITE

    current_row = 5

    # ================== ORDER INFORMATION ==================
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    ws.cell(row=current_row, column=1, value="ORDER INFORMATION")
    style_row(ws, current_row, bold=True, size=12, text_color=CL_WHITE,
              fill=FILL_INFO_HEADER, start_col=1, end_col=7,
              border=border_all, halign="center")
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    unique_id = generate_unique_order_id()
    info_rows = [
        ("ORDER DATE", order_date if order_date else "N/A"),
        ("DEALER NAME", dealer_name if dealer_name else "N/A"),
        ("CITY", city if city else "N/A"),
        ("FREIGHT", freight_condition if freight_condition else "N/A"),
        ("ORDER ID", unique_id),
    ]
    for label, value in info_rows:
        ws.cell(row=current_row, column=1, value=label)
        style_row(ws, current_row, start_col=1, end_col=1, bold=True, fill=FILL_GRAY_MED)
        ws.cell(row=current_row, column=1).border = border_all
        
        # Merge B..E for value
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
        ws.cell(row=current_row, column=2, value=value)
        fill_val = FILL_YELLOW_ID if label == "ORDER ID" else FILL_WHITE
        style_row(ws, current_row, start_col=2, end_col=5, bold=False, fill=fill_val, border=border_all, halign="center")
        for c in range(6, 8):
            ws.cell(row=current_row, column=c).fill = FILL_WHITE
            ws.cell(row=current_row, column=c).border = border_all
        current_row += 1

    current_row += 1  # spacing

    # ============ PRODUCT DETAILS & INVENTORY ============
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    ws.cell(row=current_row, column=1, value="PRODUCT DETAILS")
    style_row(ws, current_row, bold=True, size=12, text_color=CL_WHITE,
              fill=FILL_BLUE_HDR, start_col=1, end_col=7,
              border=border_all, halign="center")
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    # Table header (dark green + white)
    headers = ["PRODUCT", "SIZE", "CATEGORY", "BRAND", "QTY", "SQFT", "WEIGHT"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=i, value=h)
        cell.font = Font(bold=True, size=10, color=CL_WHITE)
        cell.fill = FILL_GREEN_DARK
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_header
    current_row += 1

    # ---------- Grouping ----------
    try:
        brand_groups = df.groupby("BRAND", sort=False)
    except Exception:
        brand_groups = [(df.iloc[0]["BRAND"] if not df.empty else "Default", df)]

    all_qty_ranges, all_sqft_ranges, all_wgt_ranges = [], [], []
    zebra = 0

    def write_brand_strip(name):
        nonlocal current_row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        ws.cell(row=current_row, column=1, value=f"Brand: {name}")
        style_row(ws, current_row, start_col=1, end_col=7, bold=True, size=11,
                  text_color=CL_BLACK, fill=FILL_ACCENT_LIGHT, border=border_all, halign="left")
        current_row += 1

    def write_category_strip(cat):
        nonlocal current_row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        ws.cell(row=current_row, column=1, value=f"CATEGORY : {cat}")
        style_row(ws, current_row, start_col=1, end_col=7, bold=True, size=10,
                  text_color=CL_BLACK, fill=FILL_ACCENT_LIGHT, border=border_all, halign="left")
        current_row += 1

    def write_category_subtotal(start_r, end_r):
        nonlocal current_row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        ws.cell(row=current_row, column=1, value="SUBTOTAL")
        ws.cell(row=current_row, column=5, value=f"=SUM(E{start_r}:E{end_r})")
        ws.cell(row=current_row, column=6, value=f"=SUM(F{start_r}:F{end_r})")
        ws.cell(row=current_row, column=7, value=f"=SUM(G{start_r}:G{end_r})")
        style_row(ws, current_row, start_col=1, end_col=7, bold=True, size=10,
                  text_color=CL_WHITE, fill=PatternFill("solid", fgColor="00A651"),
                  border=border_thick)
        current_row += 2  # ⬅️ extra blank line removed (was 2)

    for brand_name, bdf in brand_groups:
        write_brand_strip(brand_name)

        brand_products = bdf["PRODUCT"].str.lower().unique()
        has_laminate_or_liner = any(p in ["laminate", "liner"] for p in brand_products)

        try:
            cat_groups = bdf.groupby("CATEGORY_NORM", sort=False)
        except Exception:
            cat_groups = [("-", bdf)]

        brand_qty, brand_sqft, brand_wgt = [], [], []

        for cat_name, cdf in cat_groups:
            # ⬇️ Skip empty categories to avoid blank "CATEGORY :" strips
            if cdf.empty:
                continue

            write_category_strip(cat_name)

            data_start = current_row
            for _, r in cdf.iterrows():
                fill = FILL_WHITE if zebra % 2 == 0 else FILL_GRAY_LIGHT
                zebra += 1

                ws.cell(row=current_row, column=1, value=str(r["PRODUCT"]))
                ws.cell(row=current_row, column=2, value=str(r["SIZE"]))
                ws.cell(row=current_row, column=3, value=str(r["CATEGORY"]))
                ws.cell(row=current_row, column=4, value=str(r["BRAND"]))
                try:
                    qty_val = int(float(r["QUANTITY"]))
                except Exception:
                    qty_val = 0
                ws.cell(row=current_row, column=5, value=qty_val)
                ws.cell(row=current_row, column=6, value=get_sqft_formula(r, current_row))
                ws.cell(row=current_row, column=7, value=get_weight_formula(
                    r, weight_map, hdmr_map, mdf_map, ply_map, pvc_map, wpc_map, current_row
                ))

                style_row(ws, current_row, start_col=1, end_col=7, fill=fill, border=border_all)
                current_row += 1

            if current_row > data_start:
                data_end = current_row - 1
                brand_qty.append(f"E{data_start}:E{data_end}")
                brand_sqft.append(f"F{data_start}:F{data_end}")
                brand_wgt.append(f"G{data_start}:G{data_end}")
                write_category_subtotal(data_start, data_end)

        # BRAND TOTAL - Conditional inclusion (ONLY if it has laminate/liner)
        if has_laminate_or_liner:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            ws.cell(row=current_row, column=1, value="BRAND TOTAL")
            if brand_qty:
                ws.cell(row=current_row, column=5, value=f"=SUM({','.join(brand_qty)})")
                ws.cell(row=current_row, column=6, value=f"=SUM({','.join(brand_sqft)})")
                ws.cell(row=current_row, column=7, value=f"=SUM({','.join(brand_wgt)})")
            else:
                ws.cell(row=current_row, column=5, value=0)
                ws.cell(row=current_row, column=6, value=0)
                ws.cell(row=current_row, column=7, value=0)

            style_row(ws, current_row, start_col=1, end_col=7, bold=True, size=11,
                      text_color=CL_WHITE, fill=FILL_BLUE_HDR, border=border_thick, halign="center")
            current_row += 3
        else:
            # ⬇️ No extra spacer when brand total is skipped
            pass

        # Add all category totals to grand total ranges
        all_qty_ranges.extend(brand_qty)
        all_sqft_ranges.extend(brand_sqft)
        all_wgt_ranges.extend(brand_wgt)

    # --------------- GRAND TOTAL (gold) ---------------
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=1, value="GRAND TOTAL")
    ws.cell(row=current_row, column=5, value=f"=SUM({','.join(all_qty_ranges)})" if all_qty_ranges else 0)
    ws.cell(row=current_row, column=6, value=f"=SUM({','.join(all_sqft_ranges)})" if all_sqft_ranges else 0)
    ws.cell(row=current_row, column=7, value=f"=SUM({','.join(all_wgt_ranges)})" if all_wgt_ranges else 0)
    style_row(ws, current_row, start_col=1, end_col=7, bold=True, size=12,
              text_color=CL_BLACK, fill=FILL_GOLD, border=border_thick, halign="center")
    current_row += 1

    # --------------- Footer ---------------
    total_items  = int(len(df.index))
    total_brands = int(df["BRAND"].nunique()) if not df.empty else 0
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    ws.cell(row=current_row, column=1,
            value=f"ⓘ  Report Generated by NT Wood Management System | Total items: {total_items} | Brands: {total_brands}")
    style_row(ws, current_row, start_col=1, end_col=7, bold=False, size=9,
              text_color=CL_BLACK, fill=FILL_FOOTER, border=border_all, halign="center")

    # Save + Audit
    wb.save(output_file)
    try:
        log_order_to_database(username, dealer_name, city, unique_id, os.path.basename(output_file))
    except Exception:
        pass
